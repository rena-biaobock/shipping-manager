import os
import openpyxl
from datetime import date, datetime

_CONDITION_MAP = {
    "antecipa futuro": "antecipa_futuro",
    "fixo futuro": "fixo_futuro",
    "pedido até hoje": "pedido_ate_hoje",
    "pedido ate hoje": "pedido_ate_hoje",
    "fixo mês atual": "fixo_mes_atual",
    "fixo mes atual": "fixo_mes_atual",
    "antecipa mês atual": "antecipa_mes_atual",
    "antecipa mes atual": "antecipa_mes_atual",
}


_EMBARQUE_NULLS = frozenset({"0", "", "-"})


def _map_row(row: dict) -> dict:
    embarque_raw = str(row.get("Embarque Etiq") or "").strip()
    embarque_id = embarque_raw if embarque_raw not in _EMBARQUE_NULLS else None
    has_pedido = bool(row.get("Pedido"))

    if embarque_id:
        status = "in_transit_to_terminal"
    elif has_pedido:
        status = "reserved"
    else:
        status = "available_in_stock"

    exit_date = None
    raw_date = row.get("Data Saida Pedido")
    if isinstance(raw_date, (date, datetime)):
        exit_date = raw_date.strftime("%Y-%m-%d")
    elif isinstance(raw_date, str) and raw_date.strip():
        try:
            exit_date = datetime.fromisoformat(raw_date.strip().replace(" ", "T")).strftime("%Y-%m-%d")
        except ValueError:
            exit_date = None

    cond_key = str(row.get("Pedido Condição") or "").lower().strip()
    order_condition = _CONDITION_MAP.get(cond_key) or (cond_key.replace(" ", "_") if cond_key else None)

    try:
        volume_tons = float(row.get("Volume Geral") or 0) / 1000
    except (TypeError, ValueError):
        volume_tons = 0.0

    try:
        raw_pcs = str(row.get("Qt PC") or "0").split(".")[0]
        piece_count = int(raw_pcs)
    except (ValueError, AttributeError):
        piece_count = 0

    return {
        "progressivo": str(row.get("progressivo") or ""),
        "item_code": str(row.get("Item") or ""),
        "description": str(row.get("Descricao") or ""),
        "customer": str(row.get("Cliente Ped") or ""),
        "country": str(row.get("País") or ""),
        "order_number": str(row["Pedido"]) if row.get("Pedido") else None,
        "is_standard_bundle": row.get("Fardo Padrão") == "Sim",
        "embarque_id": embarque_id,
        "volume_tons": volume_tons,
        "piece_count": piece_count,
        "order_condition": order_condition,
        "exit_date": exit_date,
        "warehouse_code": str(row.get("Wharehouse") or ""),
        "status": status,
        "actual_length_m": None,
        "address": None,
        "nf": None,
        "invoice": None,
        "scan_count": 0,
        "last_scanned_at": None,
        "days_without_scan": None,
        "avg_days_idle": None,
    }


_cache: list[dict] | None = None


def load_labels() -> list[dict]:
    global _cache
    if _cache is not None:
        return _cache

    xlsx_path = os.getenv("XLSX_PATH", "/data/stock.xlsx")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.worksheets[0]

    headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
    rows = []
    for row_values in ws.iter_rows(min_row=2, values_only=True):
        row = dict(zip(headers, row_values))
        mapped = _map_row(row)
        if mapped["progressivo"]:
            rows.append(mapped)

    wb.close()
    _cache = rows
    return _cache


def invalidate_cache() -> None:
    global _cache
    _cache = None
